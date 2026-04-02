#!/usr/bin/env python3
"""Generate DSR Excel for Guidance for AI-Powered Restaurant Visibility on AWS."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True, size=10)
NORM = Font(size=10)
TITLE = Font(bold=True, size=14)
YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(wrap_text=True, vertical='top')

def hdr(ws, row, cols):
    for c, v in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical='center')

def drow(ws, row, vals):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = THIN; cell.alignment = WRAP; cell.font = NORM

# ============ SUMMARY TAB ============
ws = wb.active
ws.title = "Summary"
ws.cell(row=1, column=3, value="Guidance for AI-Powered Restaurant Visibility on AWS").font = TITLE
ws.cell(row=3, column=2, value="Project Name").font = BOLD
ws.cell(row=3, column=3, value="ai-powered-restaurant-visibility").font = NORM
ws.cell(row=4, column=2, value="Project ID").font = BOLD
ws.cell(row=5, column=2, value="Version").font = BOLD
ws.cell(row=5, column=3, value="1.0").font = NORM

summary_sections = [
    "Specify services in use", "I. General", "II. Compute", "III. Storage",
    "IV. Databases", "V. Network & Delivery", "VI. Management & Governance",
    "VII. Machine Learning", "VIII. Analytics", "IX. Security & Compliance",
    "X. Serverless", "XI. Application Integration", "XII. Media Services",
    "XIII. Developer Tools", "XIV. Internet of Things"
]
hdr(ws, 7, ["Index", "", "Identified", "Mitigated", "Not Mitigated", "Mitigation Progress", "Assessment Progress"])
for i, s in enumerate(summary_sections):
    ws.cell(row=8+i, column=1, value=s).font = NORM
    for c in range(2, 8):
        ws.cell(row=8+i, column=c).border = THIN
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15

# ============ LIST OF SERVICES TAB ============
ws2 = wb.create_sheet("List of Services")
ws2.cell(row=1, column=1, value="List of Services Used in This Project").font = TITLE
ws2.cell(row=3, column=1, value="Service").font = BOLD
ws2.cell(row=3, column=2, value="In Use?").font = BOLD
ws2.cell(row=3, column=3, value="Category").font = BOLD

services = [
    ("Amazon S3", "Yes", "III. Storage"),
    ("Amazon CloudFront", "Yes", "V. Network & Delivery"),
    ("API Gateway", "Yes", "V. Network & Delivery"),
    ("CloudFormation", "Yes", "VI. Management & Governance"),
    ("CloudWatch", "Yes", "VI. Management & Governance"),
    ("CloudTrail", "Yes", "VI. Management & Governance"),
    ("Amazon Bedrock", "Yes", "VII. Machine Learning"),
    ("Amazon Bedrock Agents", "Yes", "VII. Machine Learning"),
    ("Amazon Bedrock Knowledge Bases", "Yes", "VII. Machine Learning"),
    ("Amazon OpenSearch Serverless", "Yes", "VIII. Analytics"),
    ("IAM", "Yes", "IX. Security & Compliance"),
    ("Cognito", "Yes", "IX. Security & Compliance"),
    ("Key Management Service", "Yes", "IX. Security & Compliance"),
    ("Lambda", "Yes", "X. Serverless"),
    ("Dynamo DB", "Yes", "IV. Databases"),
    ("SQS", "Yes", "XI. Application Integration"),
    ("WAF", "Yes", "V. Network & Delivery"),
    ("EC2", "No", "II. Compute"), ("ECS", "No", "II. Compute"),
    ("EKS", "No", "II. Compute"), ("RDS", "No", "IV. Databases"),
    ("VPC", "No", "V. Network & Delivery"), ("Secrets Manager", "No", "IX. Security & Compliance"),
    ("Step Functions", "No", "X. Serverless"), ("SNS", "No", "XI. Application Integration"),
    ("EventBridge", "No", "XI. Application Integration"),
]
for i, (svc, used, cat) in enumerate(services):
    r = 4 + i
    ws2.cell(row=r, column=1, value=svc).font = NORM
    cell = ws2.cell(row=r, column=2, value=used)
    cell.font = NORM
    cell.fill = YES_FILL if used == "Yes" else NO_FILL
    ws2.cell(row=r, column=3, value=cat).font = NORM
    for c in range(1, 4):
        ws2.cell(row=r, column=c).border = THIN
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 30

# ============ CODE SCANNING TOOLS TAB ============
ws3 = wb.create_sheet("Code Scanning Tools")
ws3.cell(row=1, column=1, value="Code Scanning Tools").font = TITLE
ws3.cell(row=3, column=1, value="Scanner").font = BOLD
ws3.cell(row=3, column=2, value="Status").font = BOLD
ws3.cell(row=3, column=3, value="Findings").font = BOLD
ws3.cell(row=3, column=4, value="Notes").font = BOLD
for c in range(1, 5):
    ws3.cell(row=3, column=c).fill = HEADER_FILL
    ws3.cell(row=3, column=c).font = HEADER_FONT
    ws3.cell(row=3, column=c).border = THIN

scanners = [
    ("ASH v3.2.3 (Multi-scanner)", "COMPLETED", "All scanners PASSED", "Comprehensive scan March 12, 2026. Remediation applied March 15, 2026"),
    ("Bandit (Python SAST)", "PASSED", "0 actionable (4 low)", "random.uniform usage in data loader — not security-relevant"),
    ("cdk-nag (CloudFormation)", "PASSED", "25 original — all resolved", "Cognito MFA/AdvancedSecurity remediated. Remaining: false positives (SQS2, KMS5, CFR4) and by-design (APIG4 OPTIONS, IAM5 AgentCore)"),
    ("Checkov (IaC)", "PASSED", "8 original — all resolved", "KB S3 logging/versioning remediated. Remaining: by-design (Lambda VPC, SQS encryption, IAM wildcards)"),
    ("detect-secrets", "PASSED", "1 false positive", "GenerateSecret: false in Cognito client config is not a secret"),
    ("Semgrep (Code patterns)", "PASSED", "6 original — all resolved", "SRI integrity added to 3d-twin.html (2), unsafe-formatstring fixed in api.js (1). innerHTML with escapeHtml sanitizer (by design)"),
    ("npm-audit (JS deps)", "PASSED", "0", "No dependency vulnerabilities"),
    ("cfn-nag", "MISSING", "N/A", "Not installed — covered by cdk-nag"),
    ("Grype/Syft (Container)", "MISSING", "N/A", "No containers in project"),
]
for i, (name, status, findings, notes) in enumerate(scanners):
    r = 4 + i
    drow(ws3, r, [name, status, findings, notes])
    if "PASSED" in status:
        ws3.cell(row=r, column=2).fill = YES_FILL
    elif "FAILED" in status:
        ws3.cell(row=r, column=2).fill = NO_FILL
ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 60

# ============ HELPER FOR QUESTION TABS ============
Q_HEADERS = ["Category", "ID", "In Scope?", "Question", "Reason", "Response", "Comments", "Release Blocker", "Risk"]

def add_question_tab(name, questions):
    ws = wb.create_sheet(name)
    hdr(ws, 1, Q_HEADERS)
    for i, q in enumerate(questions):
        drow(ws, 2 + i, q)
        # Color the response
        resp_cell = ws.cell(row=2+i, column=6)
        if resp_cell.value and resp_cell.value.startswith("Yes"):
            resp_cell.fill = YES_FILL
        elif resp_cell.value and "No;" in str(resp_cell.value):
            resp_cell.fill = NO_FILL
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

# ============ I. GENERAL ============
general_qs = [
    ["PREREQ", "P1", "Yes", "Is this application following all applicable open source policies?",
     "Refer to OSPO team for open source contributions",
     "Yes", "AWS Guidance (open-source) published on GitHub under MIT-0 license.", "Yes", "Not Assessed Yet"],
    ["PREREQ", "P2", "Yes", "Have you validated open source library licenses against the approved list?",
     "Reference approved license list",
     "Yes", "All libraries (Strands SDK, boto3) validated. Dependencies in requirements.txt.", "Yes", "Not Assessed Yet"],
    ["PREREQ", "P3", "Yes", "Source code includes appropriate copyright headers and license?",
     "AWS owns copyright under standard SOWs",
     "Yes", "Source code includes LICENSE file (MIT-0). Dependencies managed via requirements.txt.", "Yes", "Not Assessed Yet"],
    ["SCOPE", "SC1", "Yes", "Solution designed to be immune from OWASP Top 10?",
     "Input validation, access control, encryption",
     "Yes", "Cognito auth, input validation on Lambda, parameterized DynamoDB queries. ASH Bandit/Semgrep confirmed.", "Yes", "Not Assessed Yet"],
    ["SCOPE", "SC2", "Yes", "No data movement between accounts/regions without approval?",
     "Legal/regulatory compliance",
     "Yes", "Solution operates within single AWS account and region (us-east-1). No cross-account data movement.", "Yes", "Not Assessed Yet"],
    ["SCOPE", "SC3", "Yes", "Solution uses secrets management for sensitive data?",
     "Use Secrets Manager or SSM SecureString",
     "Yes", "KMS CMKs encrypt DynamoDB tables and Lambda env vars. Table names/ARNs via CloudFormation env vars. No Secrets Manager used (no external credentials needed).", "Yes", "Not Assessed Yet"],
    ["SCOPE", "SC4", "Yes", "No hardcoded secrets or default keys in code?",
     "Hardcoded secrets lead to security incidents",
     "Yes", "No hardcoded secrets in application code. detect-secrets flagged 1 false positive (GenerateSecret: false). Demo credentials in deploy script noted as residual risk.", "Yes", "Not Assessed Yet"],
    ["SCOPE", "SC5", "Yes", "No sensitive data detected in log files?",
     "Do not log PII, secrets, passwords",
     "Yes", "Lambda error handlers return generic messages. No PII logged. Agent system prompt defines data boundaries.", "Yes", "Not Assessed Yet"],
    ["SCOPE", "SC6", "Yes", "Solution does not modify existing network access controls?",
     "Avoid accidental exposure",
     "Yes", "Solution creates its own serverless resources. Does not modify existing security groups, NACLs, or route tables.", "No", "Not Assessed Yet"],
    ["SCOPE", "SC7", "Yes", "X-Ray or tracing enabled for service integrations?",
     "Enable tracing for debugging",
     "No; explanation in comments", "X-Ray tracing not explicitly enabled. CloudWatch logging enabled for all Lambda functions and API Gateway.", "No", "Not Assessed Yet"],
    ["SCOPE", "SC8", "Yes", "Security code scanners used and Critical/High findings remediated?",
     "Run security scanners on code",
     "Yes", "ASH v3.2.3 with 10 scanners. Bandit PASSED. 40 original findings — Cognito MFA/AdvancedSecurity FIXED, KB S3 logging FIXED, SRI integrity FIXED, unsafe-formatstring FIXED. Remaining are false positives or by-design (documented in threat model).", "Yes", "Not Assessed Yet"],
    ["SCOPE", "SC9", "Yes", "Solution will not handle regulated data (PCI, HIPAA, GDPR)?",
     "Regulated data requires security consultant",
     "Yes", "Solution handles restaurant operational data (temperatures, inventory, staffing). No PCI/HIPAA/GDPR regulated data.", "Yes", "Not Assessed Yet"],
    ["SCOPE", "SC10", "Yes", "No binaries or container images shared as deliverable?",
     "Sharing binaries introduces security risks",
     "Yes", "No binaries or container images shared. AgentCore agent deployed via CLI from source code.", "Yes", "Not Assessed Yet"],
    ["IDEMPOTENCY", "R1", "Yes", "Rollback mechanisms for failed deployments?",
     "Leave account in known state on failure",
     "Yes", "CloudFormation provides built-in rollback. deploy-all.sh uses set -e for fail-fast.", "No", "Not Assessed Yet"],
    ["IDEMPOTENCY", "R2", "Yes", "Documentation enumerates all resources created?",
     "Avoid confusion and unexpected charges",
     "Yes", "README.md and deployment docs enumerate all resources. CloudFormation outputs display resource identifiers.", "No", "Not Assessed Yet"],
    ["IDEMPOTENCY", "R3", "Yes", "Multiple deployments don't leave unknown state?",
     "Idempotent deployments",
     "Yes", "CloudFormation deployments are idempotent with --no-fail-on-empty-changeset flag.", "No", "Not Assessed Yet"],
    ["IDEMPOTENCY", "R4", "Yes", "Concurrent deployments prevented?",
     "Avoid race conditions",
     "Yes", "CloudFormation stack locking prevents concurrent deployments.", "No", "Not Assessed Yet"],
    ["IDEMPOTENCY", "R5", "Yes", "Solution does not mutate pre-existing stack resources?",
     "Avoid unintended side effects",
     "Yes", "Solution creates all its own resources. Does not modify pre-existing stacks or resources.", "No", "Not Assessed Yet"],
    ["ENCRYPTION", "EN1", "Yes", "All data at rest encrypted?",
     "Use AWS managed or customer managed keys",
     "Yes", "DynamoDB: KMS CMK encryption on all 6 tables. S3: AES-256. Lambda env vars: separate KMS CMK. Both keys with auto rotation.", "No", "Not Assessed Yet"],
    ["ENCRYPTION", "EN2", "Yes", "Encryption in transit enabled for all services?",
     "Use TLS for all communication",
     "Yes", "CloudFront: HTTPS with TLS 1.2 minimum. S3 bucket policies enforce aws:SecureTransport. SQS DLQ policies enforce SSL-only. API Gateway uses HTTPS.", "No", "Not Assessed Yet"],
    ["ENCRYPTION", "EN3", "Yes", "VPC endpoints for serverless resources accessing VPC resources?",
     "Keep traffic off public internet",
     "Yes", "Lambda communicates with DynamoDB, S3 via AWS SDK (TLS by default). API Gateway to Lambda uses internal AWS networking.", "No", "Not Assessed Yet"],
    ["ENCRYPTION", "EN4", "Yes", "At-rest encryption meets customer needs for sensitive data?",
     "Consider role-based or column-based encryption",
     "Yes", "KMS CMK encryption used for all DynamoDB tables. Customers should evaluate if additional encryption needed for compliance.", "Yes", "Not Assessed Yet"],
]
add_question_tab("I. General", general_qs)

# ============ III. STORAGE ============
storage_qs = [
    ["Amazon S3", "S1", "Yes", "Access logging enabled on all S3 buckets?",
     "Visibility into actions against S3",
     "Yes", "Dashboard bucket has access logging to dedicated AccessLogsBucket with 90-day lifecycle. KB data bucket does not have logging (noted in ASH findings).", "No", "Not Assessed Yet"],
    ["Amazon S3", "S2", "Yes", "S3 buckets configured to block public access?",
     "Prevent unintended public exposure",
     "Yes", "All S3 buckets configured with BlockPublicAcls, BlockPublicPolicy, IgnorePublicAcls, RestrictPublicBuckets all set to true.", "No", "Not Assessed Yet"],
    ["Amazon S3", "S3", "Yes", "Customer need for custom encryption (SSE-KMS, SSE-C)?",
     "Compliance and security needs",
     "Yes", "S3 buckets use AES-256 (SSE-S3). For production, customers should evaluate CMK needs.", "Yes", "Not Assessed Yet"],
    ["Amazon S3", "S5", "Yes", "S3 as CloudFront origin uses OAC?",
     "Restrict direct S3 access",
     "Yes", "CloudFront uses Origin Access Control (OAC) with sigv4 signing. S3 bucket policy allows only CloudFront service principal.", "Yes", "Not Assessed Yet"],
    ["Amazon S3", "S6", "Yes", "Static website hosting has security headers?",
     "HSTS, CSP, X-Frame-Options",
     "Yes", "CloudFront SecurityHeadersPolicy includes HSTS, X-Frame-Options: DENY, X-Content-Type-Options, Referrer-Policy.", "No", "Not Assessed Yet"],
    ["Amazon S3", "S8", "Yes", "S3 lifecycle policies configured?",
     "Manage object lifecycle",
     "Yes", "AccessLogsBucket has 90-day lifecycle policy (DeleteOldLogs). Dashboard bucket has versioning enabled.", "No", "Not Assessed Yet"],
]
add_question_tab("III. Storage", storage_qs)

# ============ IV. DATABASES ============
db_qs = [
    ["DynamoDB", "DDB2", "Yes", "Need for client-side encryption of sensitive attributes?",
     "Protect credit card numbers, SSNs etc",
     "No; explanation in comments", "Client-side encryption not implemented. DynamoDB uses KMS CMK server-side encryption. No sensitive PII stored. Customers should evaluate for production.", "No", "Not Assessed Yet"],
    ["DynamoDB", "DDB3", "Yes", "Continuous backups (PITR) enabled?",
     "Point-in-time recovery for data protection",
     "Yes", "PITR enabled on all 6 DynamoDB tables via PointInTimeRecoverySpecification: PointInTimeRecoveryEnabled: true.", "No", "Not Assessed Yet"],
    ["DynamoDB", "DDB6", "Yes", "CloudTrail data plane events enabled?",
     "Audit CRUD operations",
     "No; explanation in comments", "CloudTrail data plane logging not explicitly enabled for DynamoDB. Customers should enable for production audit requirements.", "No", "Not Assessed Yet"],
    ["DynamoDB", "DDB7", "Yes", "AWS Config monitoring DynamoDB configuration changes?",
     "Compliance and drift detection",
     "No; explanation in comments", "AWS Config not configured. Customers should enable Config rules for production DynamoDB monitoring.", "No", "Not Assessed Yet"],
    ["DynamoDB", "DDB8", "Yes", "DynamoDB access roles use least privilege?",
     "Minimize blast radius",
     "Yes", "API Lambda role scoped to 5 specific table ARNs with GetItem/Query/Scan only. aws:RequestedRegion condition applied. KMS Decrypt permission scoped to DynamoDB encryption key ARN.", "No", "Not Assessed Yet"],
]
add_question_tab("IV. Databases", db_qs)

# ============ V. NETWORK & DELIVERY ============
net_qs = [
    ["CloudFront", "CFR1", "Yes", "Geography-aware rules to block/allow access?",
     "Geo restriction for content access",
     "No; explanation in comments", "Geo restriction not configured (RestrictionType: none). This is a demo. Customers should configure based on deployment geography.", "No", "Not Assessed Yet"],
    ["CloudFront", "CFR2", "Yes", "AWS WAF on public CloudFront distributions?",
     "Protect against application-layer attacks",
     "Yes", "AWS WAF configured with AWSManagedRulesCommonRuleSet and AWSManagedRulesKnownBadInputsRuleSet. WebACL attached to CloudFront distribution.", "No", "Not Assessed Yet"],
    ["CloudFront", "CFR3", "Yes", "Access logging enabled on CloudFront?",
     "Support incident investigation",
     "Yes", "CloudFront logging enabled to AccessLogsBucket with cloudfront-logs/ prefix. IncludeCookies: false.", "Yes", "Not Assessed Yet"],
    ["CloudFront", "CFR4", "Yes", "HTTPS required between viewers and CloudFront?",
     "Encrypt viewer connections",
     "Yes", "ViewerProtocolPolicy: redirect-to-https. MinimumProtocolVersion: TLSv1.2_2021.", "Yes", "Not Assessed Yet"],
    ["CloudFront", "CFR6", "Yes", "Origin Access Control for S3 origins?",
     "Restrict direct S3 access",
     "Yes", "OAC configured with SigningBehavior: always, SigningProtocol: sigv4. S3 bucket policy restricts to CloudFront distribution ARN.", "Yes", "Not Assessed Yet"],
    ["API Gateway", "APIG1", "Yes", "Access logging enabled on API Gateway?",
     "Record requests for investigation",
     "No; explanation in comments", "API Gateway access logging not explicitly configured in CloudFormation. CloudWatch Logs available via Lambda. Customers should enable API GW access logging for production.", "Yes", "Not Assessed Yet"],
    ["API Gateway", "APIG2", "Yes", "Request validation enabled?",
     "First-pass input validation",
     "Yes", "RequestValidator configured with ValidateRequestBody: true and ValidateRequestParameters: true.", "Yes", "Not Assessed Yet"],
    ["API Gateway", "APIG3", "Yes", "AWS WAF on public API Gateway endpoints?",
     "Prevent common web attacks",
     "No; explanation in comments", "WAF is on CloudFront (which fronts the API). API Gateway itself does not have a separate WAF. Customers should evaluate direct API GW WAF for production.", "No", "Not Assessed Yet"],
    ["API Gateway", "APIG4", "Yes", "API implements authentication?",
     "Require auth on all endpoints",
     "Yes", "All data endpoints use Cognito User Pool authorizer. Chat endpoint uses Cognito authorizer. OPTIONS methods use AuthorizationType: NONE (required for CORS preflight).", "Yes", "Not Assessed Yet"],
    ["API Gateway", "APIG7", "Yes", "API GW has access control (auth or network)?",
     "Prevent unauthorized access",
     "Yes", "Cognito authorizer on all protected endpoints. IdentitySource: method.request.header.Authorization.", "Yes", "Not Assessed Yet"],
]
add_question_tab("V. Network & Delivery", net_qs)

# ============ VI. MANAGEMENT & GOVERNANCE ============
mgmt_qs = [
    ["CloudFormation", "CFN2", "Yes", "Input parameters restricted to non-sensitive data?",
     "Store secrets in Secrets Manager",
     "Yes", "CloudFormation parameters use non-sensitive defaults (ProjectName, Environment). Sensitive values stored as KMS-encrypted Lambda env vars via CloudFormation refs.", "Yes", "Not Assessed Yet"],
    ["CloudWatch", "CW1", "Yes", "Log only non-sensitive data to CloudWatch?",
     "Redact sensitive data before logging",
     "Yes", "Lambda error handlers return generic 'Internal server error'. Agent system prompt defines data boundaries. No PII logged.", "Yes", "Not Assessed Yet"],
    ["CloudWatch", "CW2", "Yes", "CloudWatch Alarms on exceptional resource usage?",
     "Monitor and alert on anomalies",
     "No; explanation in comments", "CloudWatch Alarms not configured. Customers should implement alarms for Lambda errors, API 4xx/5xx, DynamoDB throttling for production.", "No", "Not Assessed Yet"],
    ["CloudWatch", "CW3", "Yes", "INFO/DEBUG logs disabled in production?",
     "Reduce verbose logging in prod",
     "No; explanation in comments", "Lambda logging level not explicitly restricted. Customers should set appropriate log levels for production.", "No", "Not Assessed Yet"],
]
add_question_tab("VI. Management & Governance", mgmt_qs)

# ============ VII. MACHINE LEARNING ============
ml_qs = [
    ["Amazon Bedrock", "BR1", "Yes", "Model invocation logging enabled?",
     "Monitor usage and detect anomalies",
     "No; explanation in comments", "Model invocation logging not explicitly enabled. Customers should enable Bedrock model invocation logging for production.", "No", "Not Assessed Yet"],
    ["Amazon Bedrock", "BR2", "Yes", "Filters on Bedrock model output?",
     "Compliance, abuse mitigation",
     "No; explanation in comments", "No Bedrock Guardrails configured. Agent system prompt defines boundaries. Customers MUST configure Guardrails (content filters, prompt attack detection, PII filters) for production.", "No", "Not Assessed Yet"],
    ["Amazon Bedrock", "BR3", "Yes", "RAG data restricted and isolated?",
     "Limit data uploaded to RAG",
     "Yes", "Knowledge Base S3 bucket contains only 6 equipment manuals. Bucket has public access blocked, SSL-only policy, versioning. KB role scoped to specific bucket ARN.", "Yes", "Not Assessed Yet"],
    ["Amazon Bedrock", "BR5", "Yes", "Content filters enabled for harmful categories?",
     "Baseline content filtering required",
     "No; explanation in comments", "No Bedrock Guardrails configured. Customers should configure harmful category filters for production.", "Yes", "Not Assessed Yet"],
    ["Amazon Bedrock", "BR6", "Yes", "Content filters for prompt attacks?",
     "Baseline prompt attack protection",
     "No; explanation in comments", "No prompt attack filters configured. Customers should enable Bedrock Guardrails with prompt attack detection for production.", "Yes", "Not Assessed Yet"],
    ["Amazon Bedrock", "BR8", "Yes", "Profanity filter configured?",
     "Baseline unsafe output protection",
     "No; explanation in comments", "Not configured. Customers should configure via Bedrock Guardrails for production.", "Yes", "Not Assessed Yet"],
    ["Amazon Bedrock", "BR9", "Yes", "PII filters configured?",
     "PII identification and redaction",
     "No; explanation in comments", "PII filters not configured. Customers should configure for production.", "Yes", "Not Assessed Yet"],
    ["Amazon Bedrock Agents", "BRA1", "Yes", "Agent has least privileged role?",
     "Minimum permissions for agent",
     "Yes", "AgentCore role configured in deploy-all.sh with scoped DynamoDB (GetItem/Query/Scan/PutItem/UpdateItem on restaurant-kitchen-assistant-* tables), Bedrock InvokeModel, KB Retrieve, AOSS access, and KMS Decrypt.", "Yes", "Not Assessed Yet"],
    ["Amazon Bedrock Agents", "BRA3", "Yes", "Logging configured for agent?",
     "Validate agent actions",
     "Yes", "CloudWatch Logs configured for Lambda functions. Agent implements structured logging via Python logger. AgentCore runtime provides invocation logs.", "Yes", "Not Assessed Yet"],
    ["Amazon Bedrock Agents", "BRA4", "Yes", "Agent Memory disabled if not required?",
     "Disable unnecessary features",
     "Yes", "AgentCore Memory (STM + LTM) explicitly enabled for conversation continuity. deploy-all.sh configures mode: STM_AND_LTM.", "No", "Not Assessed Yet"],
    ["Amazon Bedrock Agents", "BRA5", "Yes", "Guardrails configured for agent?",
     "Input/output filtering",
     "No; explanation in comments", "No Bedrock Guardrails configured. Customers MUST configure for production.", "Yes", "Not Assessed Yet"],
    ["Amazon Bedrock Agents", "BRA6", "Yes", "Agent tools purpose-built for use case?",
     "No shared functions across agents",
     "Yes", "9 purpose-built tools: get_restaurants, get_equipment, get_inventory, get_staffing, get_tickets, create_ticket, analyze_temperature, get_troubleshooting, search_equipment_manual.", "No", "Not Assessed Yet"],
    ["Amazon Bedrock KB", "BRKB1", "Yes", "Cross-account data source access configured?",
     "Bucket/KMS policies for cross-account",
     "Yes", "KB data source is in same account. S3 bucket policy grants Bedrock service role access. No cross-account access needed.", "No", "Not Assessed Yet"],
]
add_question_tab("VII. Machine Learning", ml_qs)

# ============ VIII. ANALYTICS (OpenSearch Serverless) ============
analytics_qs = [
    ["OpenSearch Serverless", "AOSS1", "Yes", "Encryption Security Policy configured?",
     "Determine encryption-at-rest key",
     "Yes", "AossEncPolicy configured with AWSOwnedKey: true for the rka-equipment-kb collection.", "Yes", "Not Assessed Yet"],
    ["OpenSearch Serverless", "AOSS2", "Yes", "Network security policy configured?",
     "Control collection accessibility",
     "Yes", "AossNetPolicy configured. AllowFromPublic: true for collection and dashboard access. Customers should restrict to VPC endpoints for production.", "Yes", "Not Assessed Yet"],
    ["OpenSearch Serverless", "AOSS4", "Yes", "No IAM Users for data access?",
     "Use roles, not users",
     "Yes", "Data access policy grants access to KBServiceRole ARN and account root. No individual IAM users.", "Yes", "Not Assessed Yet"],
    ["OpenSearch Serverless", "AOSS5", "Yes", "Data access policy configured?",
     "Control index/collection access",
     "Yes", "AossAccessPolicy configured with CreateIndex, ReadDocument, WriteDocument permissions scoped to the KB collection.", "Yes", "Not Assessed Yet"],
]
add_question_tab("VIII. Analytics", analytics_qs)

# ============ IX. SECURITY & COMPLIANCE ============
sec_qs = [
    ["IAM", "IAM1", "Yes", "Service roles restricted to least-privilege?",
     "Prefer AWS-managed policies",
     "Yes", "API Lambda role: 5 specific DynamoDB table ARNs, KMS key ARN, SQS DLQ ARN. Chat Lambda role: AgentCore invoke, SQS DLQ. KB role: specific S3 bucket, Titan model, AOSS.", "Yes", "Not Assessed Yet"],
    ["IAM", "IAM2", "Yes", "Custom IAM policies with least privileges?",
     "Avoid admin roles and wildcards",
     "Yes", "Custom policies use specific actions. Wildcards on AgentCore (Resource: *) documented as residual risk due to dynamic ARNs at deploy time.", "Yes", "Not Assessed Yet"],
    ["IAM", "IAM5", "Yes", "Metadata explains wildcard permission rationale?",
     "Increase readability",
     "Yes", "IAM policies include descriptive names (DynamoDBAccess, InvokeAgentCore, KBPolicy). Wildcard on AgentCore documented in threat model.", "Yes", "Not Assessed Yet"],
    ["IAM", "IAM11", "Yes", "IAM roles used instead of IAM Users?",
     "Temporary credentials, auto-rotated",
     "Yes", "Only IAM roles used. No IAM users created. Cognito handles human authentication.", "Yes", "Not Assessed Yet"],
    ["Cognito", "COG1", "Yes", "Password policy implemented?",
     "Enterprise password guidelines",
     "Yes", "MinimumLength: 12, RequireUppercase, RequireLowercase, RequireNumbers, RequireSymbols all true.", "Yes", "Not Assessed Yet"],
    ["Cognito", "COG2", "Yes", "MFA implemented?",
     "Organizational MFA policy",
     "Yes", "MFA enforced (SOFTWARE_TOKEN_MFA) with MfaConfiguration: ON. AdvancedSecurityMode: ENFORCED for adaptive authentication.", "Yes", "Not Assessed Yet"],
    ["Cognito", "COG3", "Yes", "Self-registration disabled if not needed?",
     "Disable unused registration",
     "Yes", "AllowAdminCreateUserOnly: true. Self-signup disabled.", "Yes", "Not Assessed Yet"],
    ["Cognito", "COG4", "Yes", "AdvancedSecurityMode set to ENFORCED?",
     "Extra security features",
     "Yes", "AdvancedSecurityMode: ENFORCED configured on Cognito User Pool for adaptive authentication and compromised credential detection.", "No", "Not Assessed Yet"],
    ["KMS", "KMS1", "Yes", "Customer-managed keys where access control needed?",
     "CMKs allow access control",
     "Yes", "Two KMS CMKs: DynamoDBEncryptionKey (6 tables) and LambdaEnvEncryptionKey (env vars). Both with EnableKeyRotation: true.", "No", "Not Assessed Yet"],
    ["KMS", "KMS2", "Yes", "CMK policies restricted to least-privilege?",
     "Separation of duties",
     "Yes", "DynamoDB key: Decrypt/Encrypt/GenerateDataKey for DynamoDB service. Lambda key: Decrypt/DescribeKey for Lambda service. Root account has full access.", "Yes", "Not Assessed Yet"],
]
add_question_tab("IX. Security & Compliance", sec_qs)

# ============ X. SERVERLESS ============
serverless_qs = [
    ["Lambda", "L1", "Yes", "Lambda function runtimes updated per lifecycle policies?",
     "Check LTS versions, patches",
     "Yes", "Lambda functions use Python 3.13 runtime (current supported version). See restaurant-monitoring-base-template.yaml and chat-endpoint.yaml.", "No", "Not Assessed Yet"],
    ["Lambda", "L2", "Yes", "Third party libraries include approved licenses?",
     "Open source license compliance",
     "Yes", "Dependencies managed via requirements.txt (Strands SDK, boto3). ASH Bandit/Semgrep scanned Python code. npm-audit PASSED.", "Yes", "Not Assessed Yet"],
    ["Lambda", "L3", "Yes", "Only non-sensitive data logged from Lambda?",
     "No PII, secrets in CloudWatch",
     "Yes", "Lambda error handlers return generic 'Internal server error'. No stack traces exposed. Agent system prompt defines data boundaries.", "Yes", "Not Assessed Yet"],
    ["Lambda", "L4", "Yes", "Secrets Manager/Parameter Store for Lambda env vars?",
     "Encrypt sensitive env vars",
     "Yes", "Lambda env vars contain table names and agent ARN (non-secret config). Env vars encrypted with dedicated KMS CMK (LambdaEnvEncryptionKey).", "Yes", "Not Assessed Yet"],
    ["Lambda", "L7", "Yes", "Unique IAM execution role per Lambda function?",
     "1:1 relationship Lambda to role",
     "Yes", "API Lambda has ApiLambdaRole. Chat Lambda has ChatLambdaRole. Each with separate policies.", "Yes", "Not Assessed Yet"],
    ["Lambda", "L8", "Yes", "Lambda IAM roles restricted to least-privilege?",
     "Only permissions required for task",
     "Yes", "API Lambda: read-only on 5 DynamoDB tables, KMS decrypt, SQS send, AgentCore invoke. Chat Lambda: AgentCore invoke, SQS send only.", "Yes", "Not Assessed Yet"],
    ["Lambda", "L9", "Yes", "On-failure destination for async invocations?",
     "Dead-letter queue for failures",
     "Yes", "Both Lambda functions have SQS dead letter queues configured (DeadLetterConfig). DLQ policies enforce SSL-only access.", "Yes", "Not Assessed Yet"],
]
add_question_tab("X. Serverless", serverless_qs)

# ============ XI. APPLICATION INTEGRATION ============
appint_qs = [
    ["SQS", "SQS1", "Yes", "SQS policies restricted to known IAM principals?",
     "Least privilege access",
     "Yes", "SQS DLQ policies restrict access. EnforceSSLOnly policy denies all actions when aws:SecureTransport is false.", "Yes", "Not Assessed Yet"],
    ["SQS", "SQS3", "Yes", "Dead-letter queue for unprocessable messages?",
     "Avoid message backlog",
     "Yes", "Both ApiLambdaDLQ and ChatLambdaDLQ configured as dead letter queues with 14-day message retention.", "Yes", "Not Assessed Yet"],
]
add_question_tab("XI. Application Integration", appint_qs)

# ============ SAVE ============
wb.save('temp/DSR-Filled.xlsx')
print("Done: temp/DSR-Filled.xlsx")
